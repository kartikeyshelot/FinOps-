from __future__ import annotations
import io
import time
import unittest
import pandas as pd
from openpyxl import load_workbook

from data_loader import ColumnBinding, analyze_load, finalize_binding
from instance_api import canonicalize_instance_api_name
from pricing_engine import COST_DISCLAIMER_TEXT, DEFAULT_REGION, PRICE_CACHE, get_price, get_rds_hourly
from processor import INSERT_COLS, NO_DISCOUNT, apply_na_fill, process
from rds_mysql_sa_prices import RDS_MYSQL_SA_HOURLY
from rds_recommender import get_rds_recommendations

IRELAND_EC2_LINUX_PER_HR: dict[str, float] = {
    'm5.large': 0.107,
    'm6i.large': 0.107,
    'm6g.large': 0.086,
}
IRELAND_RDS_DB_R5_LARGE: float = RDS_MYSQL_SA_HOURLY['eu-west-1']['db.r5.large']


class TestIrelandPricingLocked(unittest.TestCase):
    """EC2/RDS hourly must match embedded dataset exactly for API Name + region."""

    def test_ec2_linux_ireland(self):
        r = 'eu-west-1'
        for inst, expected in IRELAND_EC2_LINUX_PER_HR.items():
            self.assertEqual(PRICE_CACHE[r][inst], expected)
            got = get_price(inst, region=r, os='linux')
            self.assertIsNotNone(got)
            self.assertEqual(got, round(expected + 0.0, 6))

    def test_rds_db_r5_large_ireland(self):
        p = get_rds_hourly('db.r5.large', region='eu-west-1', os='linux')
        self.assertIsNotNone(p)
        self.assertEqual(p, round(IRELAND_RDS_DB_R5_LARGE, 6))

    def test_rds_db_c5_2xlarge_missing_is_na(self):
        self.assertIsNone(get_rds_hourly('db.c5.2xlarge', region='eu-west-1', os='linux'))


class TestCostFormula(unittest.TestCase):
    def test_hourly_prices_and_savings_percent(self):
        df = pd.DataFrame({'API Name': ['m5.large'], 'OS': ['linux'], 'Spend': [100.0]})
        b = ColumnBinding(instance='API Name', os='OS', actual_cost='Spend')
        out = apply_na_fill(process(df, b, region='eu-west-1', service='both', cpu_filter='both'))
        p_cur = get_price('m5.large', region='eu-west-1', os='linux')
        alt1 = out['Alt1 Instance'].iloc[0]
        self.assertEqual(alt1, 'm5a.large')
        p_alt = get_price('m5a.large', region='eu-west-1', os='linux')
        self.assertIsNotNone(p_cur)
        self.assertIsNotNone(p_alt)
        self.assertAlmostEqual(float(out['Current Price ($/hr)'].iloc[0]), float(p_cur), places=6)
        self.assertAlmostEqual(float(out['Alt1 Price ($/hr)'].iloc[0]), float(p_alt), places=6)
        if p_alt >= p_cur:
            self.assertEqual(out['Alt1 Savings %'].iloc[0], 'No Savings')
        else:
            pct = round((float(p_cur) - float(p_alt)) / float(p_cur) * 100, 1)
            self.assertEqual(float(out['Alt1 Savings %'].iloc[0]), pct)


class TestDiscountPct(unittest.TestCase):
    """Discount % = ((list hourly) - actual) / (list hourly) * 100; independent of Alt savings."""

    def test_manual_example_107_vs_096(self):
        df = pd.DataFrame({'API Name': ['m5.large'], 'OS': ['linux'], 'Spend': [0.096]})
        b = ColumnBinding(instance='API Name', os='OS', actual_cost='Spend')
        out = apply_na_fill(process(df, b, region='eu-west-1', service='both'))
        self.assertAlmostEqual(float(out['Current Price ($/hr)'].iloc[0]), 0.107, places=4)
        self.assertEqual(float(out['Actual Cost ($)'].iloc[0]), 0.096)
        self.assertEqual(float(out['Discount %'].iloc[0]), 10.3)

    def test_no_discount_when_actual_ge_list(self):
        df = pd.DataFrame({'API Name': ['m5.large'], 'OS': ['linux'], 'Spend': [0.2]})
        b = ColumnBinding(instance='API Name', os='OS', actual_cost='Spend')
        out = apply_na_fill(process(df, b, region='eu-west-1', service='both'))
        self.assertEqual(out['Discount %'].iloc[0], NO_DISCOUNT)

    def test_na_invalid_instance_no_list_price(self):
        df = pd.DataFrame({'API Name': ['not-valid'], 'OS': ['linux'], 'Spend': [0.096]})
        b = ColumnBinding(instance='API Name', os='OS', actual_cost='Spend')
        out = apply_na_fill(process(df, b, region='eu-west-1', service='both'))
        self.assertEqual(out['Discount %'].iloc[0], 'N/A')

    def test_discount_after_actual_cost_in_insert_cols(self):
        self.assertEqual(INSERT_COLS[1], 'Actual Cost ($)')
        self.assertEqual(INSERT_COLS[2], 'Discount %')
        self.assertEqual(INSERT_COLS[3], 'Current Price ($/hr)')


class TestColumnIntegrityAndInsertion(unittest.TestCase):
    def test_complex_headers_order_and_values(self):
        cols = ['C1', 'C2', 'OS', 'X', 'Instance', 'C5', 'Cost', 'C6', 'C7']
        row = ['a', 'b', 'linux', 'x', 'm5.large', 'e', 250.5, 'g', 'h']
        df = pd.DataFrame([row], columns=cols)
        b = ColumnBinding(instance='Instance', os='OS', actual_cost='Cost')
        out = process(df, b, region='eu-west-1', service='both')
        ins_idx = cols.index('Instance')
        self.assertEqual(list(out.columns[: ins_idx + 1]), cols[: ins_idx + 1])
        mid = list(out.columns[ins_idx + 1 : ins_idx + 1 + len(INSERT_COLS)])
        self.assertEqual(mid, INSERT_COLS)
        self.assertEqual(list(out.columns[ins_idx + 1 + len(INSERT_COLS) :]), cols[ins_idx + 1 :])
        for i, c in enumerate(cols):
            if i <= ins_idx:
                j = i
            else:
                j = i + len(INSERT_COLS)
            self.assertEqual(out.iloc[0, j], row[i], msg=c)


class TestServiceModeIsolation(unittest.TestCase):
    def setUp(self):
        self.df = pd.DataFrame(
            {
                'Instance': ['m5.large', 'db.r5.large'],
                'OS': ['linux', 'linux'],
                'Cost': [100.0, 200.0],
            }
        )
        self.b = ColumnBinding(instance='Instance', os='OS', actual_cost='Cost')

    def test_ec2_mode_routes_db_rows_to_rds_pipeline(self):
        """EC2 service still enriches db.* rows using RDS hourly + rds_recommender (per-row routing)."""
        out = apply_na_fill(process(self.df, self.b, region='eu-west-1', service='ec2', cpu_filter='both'))
        self.assertNotEqual(out['Alt1 Instance'].iloc[0], 'N/A')
        self.assertNotEqual(out['Alt1 Instance'].iloc[1], 'N/A')
        self.assertTrue(str(out['Alt1 Instance'].iloc[1]).startswith('db.'))

    def test_rds_only_skips_ec2_rows(self):
        out = apply_na_fill(process(self.df, self.b, region='eu-west-1', service='rds', cpu_filter='both'))
        self.assertEqual(out['Alt1 Instance'].iloc[0], 'N/A')
        self.assertNotEqual(out['Alt1 Instance'].iloc[1], 'N/A')

    def test_both_processes_each(self):
        out = apply_na_fill(process(self.df, self.b, region='eu-west-1', service='both', cpu_filter='both'))
        self.assertNotEqual(out['Alt1 Instance'].iloc[0], 'N/A')
        self.assertNotEqual(out['Alt1 Instance'].iloc[1], 'N/A')


class TestColumnDetectionSynonyms(unittest.TestCase):
    def test_vm_size_system_amount_auto_binding(self):
        df = pd.DataFrame(
            {
                'vm_size': ['m5.large'],
                'system': ['Amazon Linux 2'],
                'amount': ['123.45'],
            }
        )
        lr = analyze_load(df, [])
        self.assertFalse(lr.needs_manual_mapping)
        self.assertIsNotNone(lr.binding)
        self.assertEqual(lr.binding.instance, 'vm_size')
        self.assertEqual(lr.binding.os, 'system')
        self.assertEqual(lr.binding.actual_cost, 'amount')


class TestExtremeEdges(unittest.TestCase):
    def test_unknown_and_underscore_invalid(self):
        self.assertIsNone(canonicalize_instance_api_name('unknown.type'))
        self.assertIsNone(canonicalize_instance_api_name('m5_large'))

    def test_mixed_bad_rows_no_crash(self):
        df = pd.DataFrame(
            {
                'Instance': ['M5.LARGE', 'unknown.type', 'm5_large', 'm5.large'],
                'OS': ['linux', 'linux', 'linux', 'linux'],
                'Cost': [10.0, 20.0, 30.0, 40.0],
            }
        )
        b = ColumnBinding(instance='Instance', os='OS', actual_cost='Cost')
        out = apply_na_fill(process(df, b, region='eu-west-1', service='both'))
        self.assertEqual(len(out), 4)
        self.assertNotEqual(out['Alt1 Instance'].iloc[0], 'N/A')
        self.assertEqual(out['Alt1 Instance'].iloc[1], 'N/A')
        self.assertEqual(out['Alt1 Instance'].iloc[2], 'N/A')

    def test_missing_cost_column(self):
        df = pd.DataFrame({'Instance': ['m5.large'], 'OS': ['linux']})
        lr = finalize_binding(analyze_load(df, []), 'Instance', 'OS', None)
        out = apply_na_fill(process(lr.df, lr.binding, region='eu-west-1'))
        self.assertTrue(pd.isna(out['Actual Cost ($)'].iloc[0]))
        self.assertEqual(out['Discount %'].iloc[0], 'N/A')
        self.assertNotEqual(out['Current Price ($/hr)'].iloc[0], 'N/A')
        self.assertNotEqual(out['Alt1 Price ($/hr)'].iloc[0], 'N/A')


class TestRdsMsg1Shape(unittest.TestCase):
    """RDS reuses EC2 family map: Alt1/Alt2 imply next Intel / Graviton / latest per CPU mode."""

    def test_db_m5_intel_graviton_both(self):
        r_both = get_rds_recommendations('db.m5.large', cpu_filter='both')
        self.assertEqual(r_both['alt1'], 'db.m6g.large')
        self.assertEqual(r_both['alt2'], 'db.m7g.large')
        # No cheaper Intel-only RDS alternative exists (m6i same price, m7i more expensive).
        r_intel = get_rds_recommendations('db.m5.large', cpu_filter='intel')
        self.assertIsNone(r_intel['alt1'])
        self.assertIsNone(r_intel['alt2'])
        r_grav = get_rds_recommendations('db.m5.large', cpu_filter='graviton')
        self.assertEqual(r_grav['alt1'], 'db.m6g.large')
        self.assertEqual(r_grav['alt2'], 'db.m7g.large')


class TestExcelExportShape(unittest.TestCase):
    def test_excel_columns_match_dataframe(self):
        from excel_export import build_excel

        df = pd.DataFrame({'A': [1], 'Instance': ['m5.large'], 'B': [2], 'OS': ['linux'], 'Cost': [50.0]})
        b = ColumnBinding(instance='Instance', os='OS', actual_cost='Cost')
        proc = apply_na_fill(process(df, b, region='eu-west-1'))
        bio = io.BytesIO(build_excel(proc, 'EU (Ireland)', 'eu-west-1'))
        bio.seek(0)
        wb = load_workbook(bio, read_only=True)
        ws = wb.active
        a1 = ws['A1'].value
        self.assertIsNotNone(a1)
        self.assertIn('eu-west-1', str(a1))
        self.assertEqual(str(a1).strip(), COST_DISCLAIMER_TEXT)
        snap = ws['A2'].value
        self.assertIn('Pricing Snapshot:', str(snap))
        self.assertIn('eu-west-1', str(snap))
        hdr_row = 6
        headers = [ws.cell(row=hdr_row, column=j).value for j in range(1, len(proc.columns) + 1)]
        self.assertEqual(headers, list(proc.columns))


class TestPreshipServiceSmoke(unittest.TestCase):
    """EC2 / RDS / both enrichment: no crash, column placement, savings shape."""

    def setUp(self):
        self.df = pd.DataFrame(
            {
                'Instance': ['m5.large', 'db.r5.large'],
                'OS': ['linux', 'linux'],
                'Cost': [100.0, 200.0],
            }
        )
        self.b = ColumnBinding(instance='Instance', os='OS', actual_cost='Cost')

    def test_ec2_mode(self):
        out = apply_na_fill(process(self.df, self.b, region='eu-west-1', service='ec2', cpu_filter='both'))
        self.assertEqual(
            list(out.columns[:9]),
            [
                'Instance',
                'Pricing OS',
                'Actual Cost ($)',
                'Discount %',
                'Current Price ($/hr)',
                'Alt1 Instance',
                'Alt1 Price ($/hr)',
                'Alt1 Savings %',
                'Alt2 Instance',
            ],
        )
        self.assertNotEqual(out['Alt1 Instance'].iloc[0], 'N/A')
        self.assertNotEqual(out['Alt1 Instance'].iloc[1], 'N/A')

    def test_rds_mode(self):
        out = apply_na_fill(process(self.df, self.b, region='eu-west-1', service='rds', cpu_filter='both'))
        self.assertEqual(out['Alt1 Instance'].iloc[0], 'N/A')
        self.assertNotEqual(out['Alt1 Instance'].iloc[1], 'N/A')

    def test_both_mode(self):
        out = apply_na_fill(process(self.df, self.b, region='eu-west-1', service='both', cpu_filter='both'))
        self.assertNotEqual(out['Alt1 Instance'].iloc[0], 'N/A')
        self.assertNotEqual(out['Alt1 Instance'].iloc[1], 'N/A')

    def test_excel_matches_ui_column_order(self):
        from excel_export import build_excel

        out = apply_na_fill(process(self.df, self.b, region='eu-west-1', service='both'))
        bio = io.BytesIO(build_excel(out, 'EU (Ireland)', 'eu-west-1'))
        bio.seek(0)
        wb = load_workbook(bio, read_only=True)
        ws = wb.active
        hdr = [ws.cell(row=6, column=j).value for j in range(1, len(out.columns) + 1)]
        self.assertEqual(hdr, list(out.columns))


class TestStressMixedDataset(unittest.TestCase):
    """Messy real-world headers + mixed EC2/RDS + invalid rows — no crash."""

    def test_synonym_headers_mixed_ec2_rds_invalid(self):
        df = pd.DataFrame(
            {
                'compute_class': ['m5.large', 'db.r5.large', 'not.real', 'C7G.XLARGE'],
                'platform': ['amazon linux', 'linux', 'linux', 'Linux'],
                'billing_amount': [100.0, 200.0, 50.0, 80.0],
            }
        )
        lr = analyze_load(df, [])
        self.assertIsNotNone(lr.binding)
        out = apply_na_fill(process(lr.df, lr.binding, region='eu-west-1', service='both', cpu_filter='both'))
        self.assertEqual(len(out), 4)
        self.assertEqual(out['Alt1 Instance'].iloc[2], 'N/A')
        self.assertNotEqual(out['Alt1 Instance'].iloc[0], 'N/A')
        self.assertNotEqual(out['Alt1 Instance'].iloc[1], 'N/A')


class TestPerformance10k(unittest.TestCase):
    def test_10k_rows_completes_quickly(self):
        n = 10_500
        df = pd.DataFrame(
            {'i': ['c5.xlarge'] * n, 'o': ['linux'] * n, 'c': [1.0] * n}
        )
        b = ColumnBinding(instance='i', os='o', actual_cost='c')
        t0 = time.perf_counter()
        out = process(df, b, region=DEFAULT_REGION, service='both')
        elapsed = time.perf_counter() - t0
        self.assertEqual(len(out), n)
        self.assertLess(elapsed, 120.0, msg=f'10k+ processing took {elapsed:.1f}s')


if __name__ == '__main__':
    unittest.main()
