from __future__ import annotations

import io
import unittest

import pandas as pd
from openpyxl import load_workbook

from excel_export import build_excel


class TestExcelValidationReportSheet(unittest.TestCase):
    def test_validation_report_sheet_exists_and_has_content(self) -> None:
        df = pd.DataFrame(
            {
                "Instance": ["m5.large", "unknown.type"],
                "Current Price ($/hr)": [0.107, "N/A"],
                "Alt1 Instance": ["m6i.large", "N/A"],
                "Alt2 Instance": ["m7g.large", "N/A"],
                "Alt1 Price ($/hr)": [0.107, "N/A"],
                "Alt2 Price ($/hr)": [0.091, "N/A"],
                "Discount %": ["No Discount", "N/A"],
                "Actual Cost ($)": [0.2, pd.NA],
            }
        )
        bio = io.BytesIO(build_excel(df, "EU (Ireland)", "eu-west-1"))
        bio.seek(0)
        wb = load_workbook(bio, read_only=True)

        self.assertIn("Validation Report", wb.sheetnames)
        ws = wb["Validation Report"]
        headers = [ws.cell(row=1, column=1).value, ws.cell(row=1, column=2).value]
        self.assertEqual(headers, ["Metric", "Value"])

        # Summary rows
        self.assertEqual(ws.cell(row=2, column=1).value, "Rows (data)")
        self.assertEqual(int(ws.cell(row=2, column=2).value), 2)

        # Ensure row-level reason section exists.
        labels = [ws.cell(row=r, column=1).value for r in range(1, 20)]
        self.assertIn("Row diagnostics (below)", labels)


if __name__ == "__main__":
    unittest.main()
