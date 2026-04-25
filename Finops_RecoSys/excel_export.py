from __future__ import annotations
import io
from datetime import datetime
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from os_resolve import PRICING_OS_METADATA_NOTE
from pricing_engine import CACHE_METADATA, DECISION_SUPPORT_NOTE, PRICING_SOURCE_LABEL, _last_updated_utc, cost_disclaimer_text, format_pricing_snapshot_line

_FORMULA_TRIGGER_CHARS = frozenset('=@+-')


def sanitize_formula_injection_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Prefix string cells that could be interpreted as Excel formulas when the file is opened (= + - @)."""
    out = df.copy()
    for c in out.columns:
        s = out[c]
        if pd.api.types.is_numeric_dtype(s) and not pd.api.types.is_object_dtype(s):
            continue

        def _fix(v: object) -> object:
            if v is None:
                return v
            try:
                if pd.isna(v):
                    return v
            except (TypeError, ValueError):
                pass
            if isinstance(v, bool):
                return v
            if isinstance(v, (int, float)):
                if isinstance(v, float) and pd.isna(v):
                    return v
                return v
            if isinstance(v, bytes):
                try:
                    text = v.decode('utf-8', errors='replace')
                except Exception:
                    text = str(v)
            else:
                text = str(v)
            stripped = text.lstrip()
            if stripped and stripped[0] in _FORMULA_TRIGGER_CHARS:
                return "'" + text
            return v

        out[c] = s.map(_fix)
    return out


def savings_numeric(v) -> float | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, str):
        if v.strip() == 'No Savings':
            return 0.0
        try:
            return float(v.replace('%', ''))
        except ValueError:
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _is_na_like(v: object) -> bool:
    if v is None:
        return True
    try:
        if pd.isna(v):
            return True
    except (TypeError, ValueError):
        pass
    if isinstance(v, str):
        s = v.strip().lower()
        return s in ('', 'n/a', 'nan', 'none')
    return False


def _validation_report_df(df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame(index=df.index)
    out['row_index'] = df.index.astype(int)
    if 'Current Price ($/hr)' not in df.columns:
        out['na_reason'] = 'No current price column in output'
        return out
    out['current_price'] = df['Current Price ($/hr)']
    out['alt1_instance'] = df['Alt1 Instance'] if 'Alt1 Instance' in df.columns else pd.NA
    out['alt1_price'] = df['Alt1 Price ($/hr)'] if 'Alt1 Price ($/hr)' in df.columns else pd.NA
    out['alt2_instance'] = df['Alt2 Instance'] if 'Alt2 Instance' in df.columns else pd.NA
    out['alt2_price'] = df['Alt2 Price ($/hr)'] if 'Alt2 Price ($/hr)' in df.columns else pd.NA

    if 'Pricing OS' in df.columns:
        out['pricing_os'] = df['Pricing OS']

    # Best-effort source columns for context in diagnostics.
    region_col = next((c for c in df.columns if str(c).strip().lower() == 'region'), None)
    if region_col is not None:
        out['row_region'] = df[region_col]
    inst_col = next(
        (
            c
            for c in df.columns
            if str(c).strip().lower() in ('inst_type', 'instance', 'instance type', 'api name')
        ),
        None,
    )
    if inst_col is not None:
        out['instance_value'] = df[inst_col]

    reasons: list[str] = []
    for i in df.index:
        cur = df.at[i, 'Current Price ($/hr)']
        a1i = df.at[i, 'Alt1 Instance'] if 'Alt1 Instance' in df.columns else pd.NA
        a1p = df.at[i, 'Alt1 Price ($/hr)'] if 'Alt1 Price ($/hr)' in df.columns else pd.NA
        a2i = df.at[i, 'Alt2 Instance'] if 'Alt2 Instance' in df.columns else pd.NA
        a2p = df.at[i, 'Alt2 Price ($/hr)'] if 'Alt2 Price ($/hr)' in df.columns else pd.NA
        if _is_na_like(cur):
            reasons.append('Missing current SKU in local pricing')
        elif (not _is_na_like(a1i) and _is_na_like(a1p)) or (not _is_na_like(a2i) and _is_na_like(a2p)):
            reasons.append('Missing alternative SKU in local pricing')
        elif _is_na_like(a1i) and _is_na_like(a2i):
            reasons.append('No compatible alternative')
        else:
            reasons.append('OK')
    out['na_reason'] = reasons
    return out


def build_excel(df: pd.DataFrame, region_label: str, pricing_region_id: str) -> bytes:
    buf = io.BytesIO()
    preamble_rows = 4
    startrow = preamble_rows + 1
    safe_df = sanitize_formula_injection_dataframe(df)
    ncol = max(len(safe_df.columns), 1)
    end_letter = get_column_letter(ncol)
    snapshot = format_pricing_snapshot_line(pricing_region_id)
    disclaimer = cost_disclaimer_text(pricing_region_id)
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        safe_df.to_excel(writer, index=False, sheet_name='Recommendations', startrow=startrow)
        wb = writer.book
        ws = writer.sheets['Recommendations']
        thin = Side(style='thin', color='888888')
        bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
        disc_font = Font(size=9, bold=True)
        note_font = Font(size=9)
        disc_fill = PatternFill('solid', fgColor='FEF3C7')
        for (r, txt, font, fill) in (
            (1, disclaimer, disc_font, disc_fill),
            (2, snapshot, note_font, None),
            (3, DECISION_SUPPORT_NOTE, note_font, None),
        ):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
            c = ws.cell(row=r, column=1)
            c.value = txt
            c.font = font
            if fill is not None:
                c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical='center')
            c.border = bdr
        hdr_row = startrow + 1
        hdr_fill = PatternFill('solid', fgColor='E5E7EB')
        hdr_font = Font(bold=True, color='111827', size=10)
        for cidx in range(1, ncol + 1):
            c = ws.cell(row=hdr_row, column=cidx)
            c.font = hdr_font
            c.fill = hdr_fill
            c.border = bdr
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sav_cols = [c for c in safe_df.columns if 'Savings %' in c or c == 'Discount %']
        price_cols = [c for c in safe_df.columns if ('Cost ($)' in c or '$/hr' in c)]
        green_fill = PatternFill('solid', fgColor='D1FAE5')
        amber_fill = PatternFill('solid', fgColor='FEF3C7')
        red_fill = PatternFill('solid', fgColor='FEE2E2')
        col_list = list(safe_df.columns)
        first_data_row = hdr_row + 1
        for ridx in range(first_data_row, ws.max_row + 1):
            for cidx in range(1, ncol + 1):
                cell = ws.cell(row=ridx, column=cidx)
                cell.border = bdr
                cell.alignment = Alignment(vertical='center')
                cell.font = Font(size=10)
            for name in sav_cols:
                if name not in col_list:
                    continue
                ci = col_list.index(name) + 1
                sc = ws.cell(row=ridx, column=ci)
                val = sc.value
                nv = savings_numeric(val)
                if nv is None:
                    continue
                if nv >= 20:
                    sc.fill = green_fill
                elif nv > 0:
                    sc.fill = amber_fill
                else:
                    sc.fill = red_fill
            for name in price_cols:
                if name not in col_list:
                    continue
                ci = col_list.index(name) + 1
                ws.cell(row=ridx, column=ci).number_format = '$#,##0.0000'
        for cidx in range(1, ncol + 1):
            letter = get_column_letter(cidx)
            w = max((len(str(ws.cell(row=r, column=cidx).value or '')) for r in range(1, ws.max_row + 1)), default=8)
            ws.column_dimensions[letter].width = min(w + 2, 48)
        ws.freeze_panes = ws.cell(row=first_data_row, column=1).coordinate
        ws.auto_filter.ref = f'A{hdr_row}:{end_letter}{ws.max_row}'
        ws_m = wb.create_sheet('Metadata')
        ws_m.append(['Field', 'Value'])
        ws_m.append(['Disclaimer', disclaimer])
        ws_m.append(['Pricing snapshot', snapshot])
        ws_m.append(['Decision support note', DECISION_SUPPORT_NOTE])
        ws_m.append(['Generated at', datetime.now().strftime('%Y-%m-%d %H:%M')])
        ws_m.append(['Pricing region (label)', region_label])
        ws_m.append(['Pricing region (id)', pricing_region_id])
        ws_m.append(['Pricing source', PRICING_SOURCE_LABEL])
        ws_m.append(['Dataset as-of', _last_updated_utc().strftime('%Y-%m-%d')])
        ws_m.append(['Pricing OS handling', PRICING_OS_METADATA_NOTE])
        ws_m.append(['Rows (data)', len(safe_df)])
        # Validation report: quick run-health summary + row-level reason labels.
        report = _validation_report_df(safe_df)
        ws_v = wb.create_sheet('Validation Report')
        ws_v.append(['Metric', 'Value'])
        ws_v.append(['Rows (data)', len(safe_df)])
        if 'Current Price ($/hr)' in safe_df.columns:
            cur_na = int(safe_df['Current Price ($/hr)'].map(_is_na_like).sum())
            ws_v.append(['Rows with Current Price = N/A', cur_na])
        if 'Alt1 Instance' in safe_df.columns:
            a1_na = int(safe_df['Alt1 Instance'].map(_is_na_like).sum())
            ws_v.append(['Rows with Alt1 Instance = N/A', a1_na])
        if 'Alt2 Instance' in safe_df.columns:
            a2_na = int(safe_df['Alt2 Instance'].map(_is_na_like).sum())
            ws_v.append(['Rows with Alt2 Instance = N/A', a2_na])
        ws_v.append([])
        ws_v.append(['Top reason', 'Count'])
        reason_counts = report['na_reason'].value_counts(dropna=False).to_dict()
        for reason, cnt in reason_counts.items():
            ws_v.append([str(reason), int(cnt)])
        ws_v.append([])
        ws_v.append(['Row diagnostics (below)'])
        ws_v.append(list(report.columns))
        for row in report.itertuples(index=False, name=None):
            ws_v.append(list(row))
    return buf.getvalue()
